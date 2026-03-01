from datetime import date
from pathlib import Path
from typing import List, Tuple
from copy import copy as shallow_copy

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    load_workbook = None  # type: ignore
    Workbook = None  # type: ignore
    get_column_letter = None  # type: ignore


MONTH_NAMES_RU = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь",
}


def months_between(start: date, end: date) -> List[Tuple[int, int]]:
    start_m = date(start.year, start.month, 1)
    end_m = date(end.year, end.month, 1)
    months: List[Tuple[int, int]] = []
    y, m = start_m.year, start_m.month
    while True:
        months.append((y, m))
        if y == end_m.year and m == end_m.month:
            break
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months


def ensure_openpyxl():
    global load_workbook, get_column_letter
    if load_workbook is None:
        raise RuntimeError(
            "Библиотека openpyxl не установлена. Установите её: python -m pip install openpyxl"
        )


def copy_cell_style(src_cell, dst_cell):
    dst_cell.font = shallow_copy(src_cell.font)
    dst_cell.fill = shallow_copy(src_cell.fill)
    dst_cell.border = shallow_copy(src_cell.border)
    dst_cell.alignment = shallow_copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = shallow_copy(src_cell.protection)


def process_body_template(template_path: Path, output_path: Path, months: List[Tuple[int, int]]):
    ensure_openpyxl()
    wb = load_workbook(filename=str(template_path), data_only=False)
    ws = wb.active

    base_col_idx = 4
    insert_count = max(0, len(months) - 1)

    if insert_count > 0:
        ws.insert_cols(base_col_idx + 1, amount=insert_count)

    width_d = ws.column_dimensions["D"].width
    max_row = ws.max_row

    for i in range(1, insert_count + 1):
        col_letter = get_column_letter(base_col_idx + i)
        ws.column_dimensions[col_letter].width = width_d
        for r in range(1, max_row + 1):
            src = ws.cell(row=r, column=base_col_idx)
            dst = ws.cell(row=r, column=base_col_idx + i)
            copy_cell_style(src, dst)

    for idx, (_, m) in enumerate(months):
        target_col = base_col_idx + idx
        ws.cell(row=2, column=target_col, value=MONTH_NAMES_RU[m])

    for idx, (y, _) in enumerate(months):
        if idx == 0 or y != months[idx - 1][0]:
            target_col = base_col_idx + idx
            ws.cell(row=1, column=target_col, value=str(y))

    wb.save(str(output_path))


def _read_resources(resources_path: Path) -> List[str]:
    if not resources_path.exists():
        alt = Path("resources.txt")
        if alt.exists():
            resources_path = alt
        else:
            return []
    names: List[str] = []
    content = None
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            content = resources_path.read_text(encoding=enc)
            break
        except Exception:
            pass
    if content is None:
        content = resources_path.read_text(errors="ignore")
    for line in content.splitlines():
        s = line.strip()
        if s:
            names.append(s)
    if not names:
        alt = Path("resources.txt")
        if alt.exists() and alt != resources_path:
            try:
                content = alt.read_text(encoding="utf-8-sig")
            except Exception:
                try:
                    content = alt.read_text(encoding="utf-8")
                except Exception:
                    content = alt.read_text(errors="ignore")
            for line in content.splitlines():
                s = line.strip()
                if s:
                    names.append(s)
    return names


def _copy_row(ws, src_row: int, dst_row: int, max_col: int):
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        dst.value = src.value
        copy_cell_style(src, dst)


def expand_rows_with_resources(output_path: Path, resources_path: Path):
    ensure_openpyxl()
    wb = load_workbook(filename=str(output_path), data_only=False)
    ws = wb.active

    names = _read_resources(resources_path)
    if not names:
        wb.save(str(output_path))
        return

    blocks_needed = len(names)
    insert_pairs = max(0, blocks_needed - 1)
    if insert_pairs > 0:
        for i in range(insert_pairs):
            insert_at = 5 + i * 2
            ws.insert_rows(insert_at, amount=2)

    max_col = ws.max_column
    for i in range(1, blocks_needed):
        dst_row1 = 3 + i * 2
        dst_row2 = dst_row1 + 1
        _copy_row(ws, 3, dst_row1, max_col)
        _copy_row(ws, 4, dst_row2, max_col)

    for idx, name in enumerate(names):
        target_row = 3 + idx * 2
        target_col = 2
        tl_row, tl_col = target_row, target_col
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= target_row <= rng.max_row and rng.min_col <= target_col <= rng.max_col:
                tl_row, tl_col = rng.min_row, rng.min_col
                break
        ws.cell(row=tl_row, column=tl_col, value=name)

    wb.save(str(output_path))


def expand_rows_with_resources_ws(ws, resources_path: Path):
    names = _read_resources(resources_path)
    if not names:
        return
    blocks_needed = len(names)
    insert_pairs = max(0, blocks_needed - 1)
    if insert_pairs > 0:
        for i in range(insert_pairs):
            insert_at = 5 + i * 2
            ws.insert_rows(insert_at, amount=2)
    max_col = ws.max_column
    for i in range(1, blocks_needed):
        dst_row1 = 3 + i * 2
        dst_row2 = dst_row1 + 1
        _copy_row(ws, 3, dst_row1, max_col)
        _copy_row(ws, 4, dst_row2, max_col)
    for idx, name in enumerate(names):
        target_row = 3 + idx * 2
        target_col = 2
        tl_row, tl_col = target_row, target_col
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= target_row <= rng.max_row and rng.min_col <= target_col <= rng.max_col:
                tl_row, tl_col = rng.min_row, rng.min_col
                break
        ws.cell(row=tl_row, column=tl_col, value=name)


def build_body_workbook(template_path: Path, months: List[Tuple[int, int]], resources_path: Path):
    ensure_openpyxl()
    wb = load_workbook(filename=str(template_path), data_only=False)
    ws = wb.active
    base_col_idx = 4
    insert_count = max(0, len(months) - 1)
    if insert_count > 0:
        ws.insert_cols(base_col_idx + 1, amount=insert_count)
    width_d = ws.column_dimensions["D"].width
    max_row = ws.max_row
    for i in range(1, insert_count + 1):
        col_letter = get_column_letter(base_col_idx + i)
        ws.column_dimensions[col_letter].width = width_d
        for r in range(1, max_row + 1):
            src = ws.cell(row=r, column=base_col_idx)
            dst = ws.cell(row=r, column=base_col_idx + i)
            copy_cell_style(src, dst)
    for idx, (_, m) in enumerate(months):
        target_col = base_col_idx + idx
        ws.cell(row=2, column=target_col, value=MONTH_NAMES_RU[m])
    for idx, (y, _) in enumerate(months):
        if idx == 0 or y != months[idx - 1][0]:
            target_col = base_col_idx + idx
            ws.cell(row=1, column=target_col, value=str(y))
    expand_rows_with_resources_ws(ws, resources_path)
    return wb
