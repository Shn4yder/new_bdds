from pathlib import Path
from typing import List, Tuple
from copy import copy as shallow_copy

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    load_workbook = None  # type: ignore
    Workbook = None  # type: ignore
    get_column_letter = None  # type: ignore


def ensure_openpyxl():
    global load_workbook, get_column_letter
    if load_workbook is None:
        raise RuntimeError(
            "Библиотека openpyxl не установлена. Установите её: python -m pip install openpyxl"
        )


def copy_cell_style(src_cell, dst_cell):
    dst_cell.font = shallow_copy(src_cell.font)
    dst_cell.fill = shallow_copy(src_cell.fill)
    dst_cell.border = shallow_copy(src_cell.border)
    dst_cell.alignment = shallow_copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = shallow_copy(src_cell.protection)


def process_footer_template(template_path: Path, output_path: Path, months: List[Tuple[int, int]]):
    ensure_openpyxl()
    wb = load_workbook(filename=str(template_path), data_only=False)
    ws = wb.active

    base_col_idx = 4
    insert_count = max(0, len(months) - 1)
    if insert_count > 0:
        ws.insert_cols(base_col_idx + 1, amount=insert_count)

    width_d = ws.column_dimensions["D"].width
    max_row = ws.max_row
    for i in range(1, insert_count + 1):
        col_letter = get_column_letter(base_col_idx + i)
        ws.column_dimensions[col_letter].width = width_d
        for r in range(1, max_row + 1):
            src = ws.cell(row=r, column=base_col_idx)
            dst = ws.cell(row=r, column=base_col_idx + i)
            copy_cell_style(src, dst)

    last_col = 4 + len(months)
    ranges = list(ws.merged_cells.ranges)
    for rng in ranges:
        if rng.min_row == rng.max_row and (rng.min_col >= base_col_idx or rng.max_col >= base_col_idx):
            ws.unmerge_cells(str(rng))
            ws.merge_cells(
                start_row=rng.min_row,
                start_column=rng.min_col,
                end_row=rng.max_row,
                end_column=last_col,
            )

    from openpyxl.utils import get_column_letter as _gcl
    print(f"Последний столбец подвала: {_gcl(last_col)}")
    wb.save(str(output_path))


def build_footer_workbook(template_path: Path, months: List[Tuple[int, int]]):
    ensure_openpyxl()
    wb = load_workbook(filename=str(template_path), data_only=False)
    ws = wb.active

    base_col_idx = 4
    insert_count = max(0, len(months) - 1)
    if insert_count > 0:
        ws.insert_cols(base_col_idx + 1, amount=insert_count)

    width_d = ws.column_dimensions["D"].width
    max_row = ws.max_row
    for i in range(1, insert_count + 1):
        col_letter = get_column_letter(base_col_idx + i)
        ws.column_dimensions[col_letter].width = width_d
        for r in range(1, max_row + 1):
            src = ws.cell(row=r, column=base_col_idx)
            dst = ws.cell(row=r, column=base_col_idx + i)
            copy_cell_style(src, dst)

    last_col = 4 + len(months)
    ranges = list(ws.merged_cells.ranges)
    for rng in ranges:
        if rng.min_row == rng.max_row and (rng.min_col >= base_col_idx or rng.max_col >= base_col_idx):
            ws.unmerge_cells(str(rng))
            ws.merge_cells(
                start_row=rng.min_row,
                start_column=rng.min_col,
                end_row=rng.max_row,
                end_column=last_col,
            )

    return wb
