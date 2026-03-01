from copy import copy as shallow_copy

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
except ImportError:
    load_workbook = None  # type: ignore
    Workbook = None  # type: ignore
    get_column_letter = None  # type: ignore
    Alignment = None  # type: ignore


def ensure_openpyxl():
    if load_workbook is None or Workbook is None:
        raise RuntimeError(
            "Библиотека openpyxl не установлена. Установите её: python -m pip install openpyxl"
        )


def copy_cell_style(src_cell, dst_cell):
    dst_cell.font = shallow_copy(src_cell.font)
    dst_cell.fill = shallow_copy(src_cell.fill)
    dst_cell.border = shallow_copy(src_cell.border)
    dst_cell.alignment = shallow_copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = shallow_copy(src_cell.protection)


def normalize_name(s: str) -> str:
    return (" ".join(str(s).split())).strip()
